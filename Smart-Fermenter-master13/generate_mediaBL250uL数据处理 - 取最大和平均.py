import pandas as pd
import re
from openpyxl import load_workbook

# 读取Sheet3并构建OD值映射字典
sheet3_df = pd.read_excel('培养基数据1_扩展.xlsx', sheet_name='Sheet3', header=None)
od_mapping = {}
for i in range(sheet3_df.shape[0]):
    group_letter = chr(ord('A') + i)
    for j in range(sheet3_df.shape[1]):
        group_id = f"{group_letter}{j+1}"
        od_mapping[group_id] = sheet3_df.iloc[i, j]

# 读取并处理Sheet1数据
sheet1_df = pd.read_excel('培养基数据1_扩展.xlsx', sheet_name='Sheet1')
sheet1_df['OD值'] = sheet1_df['组编号'].map(od_mapping)

# 自定义排序函数
def get_sort_key(group_id):
    match = re.match(r"([A-Z]+)(\d+)", group_id)
    return (match.group(1), int(match.group(2))) if match else (group_id, 0)

sheet1_df['sort_key'] = sheet1_df['组编号'].apply(get_sort_key)
sheet1_sorted = sheet1_df.sort_values(by='sort_key').drop(columns=['sort_key'])

# 处理数据，生成新组编号
processed_rows = []
group_counter = 1
for i in range(0, len(sheet1_sorted), 2):
    row1 = sheet1_sorted.iloc[i]
    row2 = sheet1_sorted.iloc[i+1]
    selected = row1 if row1['OD值'] > row2['OD值'] else row2
    new_row = selected.copy()
    new_row['组编号'] = str(group_counter)
    processed_rows.append(new_row.to_dict())
processed_df = pd.DataFrame(processed_rows, columns=sheet1_sorted.columns)

# ----- 新增功能：合并重复行并取OD平均值 -----
rows_to_delete = set()  # 记录需要删除的行索引
for i in range(len(processed_df)):
    if i in rows_to_delete:
        continue  # 跳过已标记删除的行
    current_row = processed_df.iloc[i]
    # 提取比较列（排除组编号和OD值）
    compare_cols = processed_df.columns[1:-1]
    for j in range(i + 1, len(processed_df)):
        if j in rows_to_delete:
            continue
        next_row = processed_df.iloc[j]
        # 判断除首尾列外是否完全相同
        if current_row[compare_cols].equals(next_row[compare_cols]):
            # 计算OD平均值并更新当前行
            avg_od = (current_row['OD值'] + next_row['OD值']) / 2
            processed_df.at[i, 'OD值'] = avg_od
            # 标记重复行待删除
            rows_to_delete.add(j)
# 删除所有标记的行并重置索引
processed_df = processed_df.drop(index=rows_to_delete).reset_index(drop=True)
# 重新生成组编号
processed_df['组编号'] = [str(i+1) for i in range(len(processed_df))]

# 写入新Sheet4
book = load_workbook('培养基数据1_扩展.xlsx')
if 'Sheet4' in book.sheetnames:
    del book['Sheet4']
with pd.ExcelWriter('培养基数据1_扩展.xlsx', engine='openpyxl', mode='a') as writer:
    processed_df.to_excel(writer, sheet_name='Sheet4', index=False)