import random
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# 定义培养基成分及其取值范围
components = {
    "酵母浸粉": (80, 130),
    "蛋白胨": (80, 130),
    "葡萄糖": (80, 130),
    "(NH4)2SO4": (80, 130),
    "K2HPO4·3H2O": (80, 130),
    "柠檬酸·H2O": (80, 130),
    "硫酸镁": (80, 130),
    "微量元素": (80, 130)
}

# 修改后的随机生成函数：每行调整1-2个随机成分
def generate_random_media(num_samples=96):
    data = []
    components_list = list(components.keys())
    
    for i in range(1, num_samples + 1):
        sample = {"组编号": i}
        # 随机决定调整1或2个成分
        num_changes = random.randint(1, 2)
        selected_components = random.sample(components_list, num_changes)
        
        for comp in components:
            if comp in selected_components:
                min_v, max_v = components[comp]
                if max_v - min_v > 5:
                    value = random.choice(list(range(min_v, max_v+1, 10)))
                else:
                    value = round(random.choice([x*0.25 for x in 
                        range(int(min_v*4), int(max_v*4)+1)]), 2)
                sample[comp] = value
            else:
                sample[comp] = 100  # 非选中成分固定为100
        data.append(sample)
    return data

# 补水量计算逻辑保持不变
def calculate_water_supplement(data):
    for sample in data:
        total = sum(sample[comp] for comp in components)
        sample["补水量"] = 1000 - total  # 总体积固定为1000ml
    return data

# 文件备份函数保持不变
def backup_existing_file(file_name):
    if os.path.exists(file_name):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file_name = file_name.replace(".xlsx", f"_backup_{timestamp}.xlsx")
        os.rename(file_name, backup_file_name)
        print(f"原文件已备份为 {backup_file_name}")

# Excel保存函数保持不变
def save_to_excel_with_sheets(data, file_name="培养基数据.xlsx"):
    backup_existing_file(file_name)
    
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        # 原始数据写入Sheet1
        df_original = pd.DataFrame(data)
        df_original.to_excel(writer, sheet_name="Sheet1", index=False)
        
        # 创建美化后的Sheet2
        wb = writer.book
        ws = wb.create_sheet("Sheet2")
        
        # 表格样式配置
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        
        # 生成所有成分分布表
        current_row = 1
        columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
        col_headers = [str(i) for i in range(1, 13)]
        
        for comp in list(components.keys()) + ["补水量"]:
            # 添加标题行并合并单元格
            ws.merge_cells(start_row=current_row, start_column=1, 
                         end_row=current_row, end_column=13)
            title_cell = ws.cell(row=current_row, column=1, value=f"{comp}分布表")
            title_cell.font = Font(bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = Alignment(horizontal="center")
            current_row += 1
            
            # 添加行列标签
            ws.cell(row=current_row, column=1, value="").fill = header_fill
            for col_num, col in enumerate(col_headers, 2):
                cell = ws.cell(row=current_row, column=col_num, value=col)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row_num, row_name in enumerate(columns, current_row+1):
                ws.cell(row=row_num, column=1, value=row_name).fill = header_fill
                ws.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF")
            
            current_row += 1
            
            # 填充数据并应用样式
            for row_idx, row_name in enumerate(columns):
                for col_idx, col_name in enumerate(col_headers):
                    group = f"{row_name}{col_name}"
                    sample = next((s for s in data if s["组编号"] == group), None)
                    cell = ws.cell(row=current_row + row_idx, column=2 + col_idx)
                    if sample:
                        cell.value = sample[comp]
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
            
            # 应用颜色渐变条件格式
            ws.conditional_formatting.add(
                f"B{current_row}:M{current_row + 8}",
                ColorScaleRule(start_type="min", start_color="FF92D050",
                               mid_type="percentile", mid_value=50, mid_color="FFFFC000",
                               end_type="max", end_color="FFFF0000")
            )
            
            current_row += 9  # 8行数据 + 1空行
            ws.cell(row=current_row, column=1).value = ""  # 分隔空行
            current_row += 1

    print(f"数据已成功保存并美化到 {file_name}")

# 其他辅助函数保持不变
def sort_and_update_group_numbers(data):
    df = pd.DataFrame(data)
    sorted_df = df.sort_values(by=list(components.keys()))
    sorted_df["组编号"] = range(1, len(sorted_df) + 1)
    return sorted_df.to_dict(orient="records")

def reassign_group_numbers(data):
    new_data = []
    for i, sample in enumerate(data):
        column = i // 12
        row = i % 12 + 1
        new_sample = sample.copy()
        new_sample["组编号"] = f"{chr(65 + column)}{row}"
        new_data.append(new_sample)
    return new_data

def duplicate_and_renumber(data):
    duplicated = []
    for item in data:
        duplicated.extend([item.copy(), item.copy()])
    
    new_data = []
    for i, sample in enumerate(duplicated):
        column = i // 12
        row = i % 12 + 1
        new_sample = sample.copy()
        new_sample["组编号"] = f"{chr(65 + column)}{row}"
        new_data.append(new_sample)
    return new_data

if __name__ == "__main__":
    media_data = generate_random_media(96)
    sorted_data = sort_and_update_group_numbers(media_data)
    
    part1 = sorted_data[:48]
    part2 = sorted_data[48:]
    
    part1_renum = reassign_group_numbers(part1)
    part2_renum = reassign_group_numbers(part2)
    
    part1_dup = duplicate_and_renumber(part1_renum)
    part2_dup = duplicate_and_renumber(part2_renum)
    
    part1_water = calculate_water_supplement(part1_dup)
    part2_water = calculate_water_supplement(part2_dup)
    
    save_to_excel_with_sheets(part1_water, "培养基数据1_扩展.xlsx")
    save_to_excel_with_sheets(part2_water, "培养基数据2_扩展.xlsx")