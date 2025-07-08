import random
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# 定义培养基成分及其取值范围
components = {
    "酵母浸粉": (40, 100),
    "蛋白胨": (40, 100),
    "葡萄糖": (40, 100),
    "(NH4)2SO4": (40, 100),
    "K2HPO4·3H2O": (40, 100),
    "柠檬酸·H2O": (40, 100),
    "硫酸镁": (40, 100),
    "微量元素": (40, 100)
}

# 随机生成96组培养基数据，按照新的取值间隔规则
def generate_random_media(num_samples=96):
    data = []
    for i in range(1, num_samples + 1):
        sample = {"组编号": i}
        for component, (min_val, max_val) in components.items():
            if max_val - min_val > 5:  # 最大范围超过5
                # 以10为间隔生成随机值
                value = random.choice([x for x in range(min_val, max_val + 1, 10)])
            else:  # 最大范围小于等于5
                # 以0.25为间隔生成随机值
                value = round(random.choice([x * 0.25 for x in range(int(min_val * 4), int(max_val * 4) + 1)]), 2)
            sample[component] = value
        data.append(sample)
    return data

# 对数据进行排序并更新组编号
def sort_and_update_group_numbers(data):
    # 将数据转换为DataFrame以便排序
    df = pd.DataFrame(data)
    
    # 按照所有列依次排序
    sorted_df = df.sort_values(by=list(components.keys()))
    
    # 更新组编号
    sorted_df["组编号"] = range(1, len(sorted_df) + 1)
    
    # 返回排序后的数据
    return sorted_df.to_dict(orient="records")

# 检查文件是否存在，若存在则备份
def backup_existing_file(file_name):
    if os.path.exists(file_name):
        # 获取当前时间戳
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 创建备份文件名
        backup_file_name = file_name.replace(".xlsx", f"_backup_{timestamp}.xlsx")
        # 备份原文件
        os.rename(file_name, backup_file_name)
        print(f"原文件已备份为 {backup_file_name}")

# 将数据保存到Excel文件，并美化Sheet2
def save_to_excel_with_sheets(data, file_name="培养基数据.xlsx"):
    backup_existing_file(file_name)
    
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        # 原始数据写入Sheet1
        df_original = pd.DataFrame(data)
        df_original.to_excel(writer, sheet_name="Sheet1", index=False)
        
        # 创建美化后的Sheet2
        wb = writer.book
        ws = wb.create_sheet("Sheet2")
        
        # 表格样式配置
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        
        # 生成所有成分分布表
        current_row = 1
        columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
        col_headers = [str(i) for i in range(1, 13)]
        
        for comp in list(components.keys()) + ["补水量"]:
            # 添加标题行并合并单元格
            ws.merge_cells(start_row=current_row, start_column=1, 
                         end_row=current_row, end_column=13)
            title_cell = ws.cell(row=current_row, column=1, value=f"{comp}分布表")
            title_cell.font = Font(bold=True, color="FFFFFF")
            title_cell.fill = header_fill
            title_cell.alignment = Alignment(horizontal="center")
            current_row += 1
            
            # 添加行列标签
            ws.cell(row=current_row, column=1, value="").fill = header_fill
            for col_num, col in enumerate(col_headers, 2):
                cell = ws.cell(row=current_row, column=col_num, value=col)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row_num, row_name in enumerate(columns, current_row+1):
                ws.cell(row=row_num, column=1, value=row_name).fill = header_fill
                ws.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF")
            
            current_row += 1
            
            # 填充数据并应用样式
            for row_idx, row_name in enumerate(columns):
                for col_idx, col_name in enumerate(col_headers):
                    group = f"{row_name}{col_name}"
                    sample = next((s for s in data if s["组编号"] == group), None)
                    cell = ws.cell(row=current_row + row_idx, column=2 + col_idx)
                    if sample:
                        cell.value = sample[comp]
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
            
            # 应用颜色渐变条件格式
            ws.conditional_formatting.add(
                f"B{current_row}:M{current_row + 8}",
                ColorScaleRule(start_type="min", start_color="FF92D050",
                               mid_type="percentile", mid_value=50, mid_color="FFFFC000",
                               end_type="max", end_color="FFFF0000")
            )
            
            current_row += 9  # 8行数据 + 1空行
            ws.cell(row=current_row, column=1).value = ""  # 分隔空行
            current_row += 1

    print(f"数据已成功保存并美化到 {file_name}")

# 计算补水量
def calculate_water_supplement(data):
    num_components = len(components)  # 成分种类数
    for sample in data:
        # 计算该行所有成分值之和
        total_component_value = sum(sample[component] for component in components)
        # 计算补水量
        water_supplement = num_components * 100 - total_component_value + 100
        # 添加补水量列
        sample["补水量"] = water_supplement
    return data

# 重新编号规则
def reassign_group_numbers(data):
    new_data = []
    rows_per_column = 12  # 每列12个编号
    columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
    
    for i, sample in enumerate(data):
        # 计算行号和列号
        column_index = i // rows_per_column
        row_index = i % rows_per_column + 1
        
        # 构造新的编号
        new_sample = sample.copy()
        new_sample["组编号"] = f"{columns[column_index]}{row_index}"
        new_data.append(new_sample)
    
    return new_data

# 新增功能：将48行数据复制为96行数据，并重新编号
def duplicate_and_renumber(data):
    duplicated_data = []
    rows_per_column = 12  # 每列12个编号
    columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
    
    # 将48行数据复制为96行数据，每行重复一次
    for i in range(len(data)):
        duplicated_data.append(data[i])  # 添加原始行
        duplicated_data.append(data[i])  # 再次添加原始行（复制）
    
    # 重新编号
    new_data = []
    for i, sample in enumerate(duplicated_data):
        # 计算行号和列号
        column_index = i // rows_per_column
        row_index = i % rows_per_column + 1
        
        # 构造新的编号
        new_sample = sample.copy()
        new_sample["组编号"] = f"{columns[column_index]}{row_index}"
        new_data.append(new_sample)
    
    return new_data

# 主函数
if __name__ == "__main__":
    # 生成100组随机培养基数据
    media_data = generate_random_media(96)
    
    # 对数据进行排序并更新组编号
    sorted_media_data = sort_and_update_group_numbers(media_data)
    
    # 将96组数据分为两部分，每部分48组
    part1 = sorted_media_data[:48]
    part2 = sorted_media_data[48:]
    
    # 重新编号
    part1_renumbered = reassign_group_numbers(part1)
    part2_renumbered = reassign_group_numbers(part2)
    
    # 将48行数据复制为96行数据，并重新编号
    part1_duplicated = duplicate_and_renumber(part1_renumbered)
    part2_duplicated = duplicate_and_renumber(part2_renumbered)
    
    # 计算补水量
    part1_with_water = calculate_water_supplement(part1_duplicated)
    part2_with_water = calculate_water_supplement(part2_duplicated)
    
    # 保存到两个Excel文件
    save_to_excel_with_sheets(part1_with_water, "培养基数据1_扩展.xlsx")
    save_to_excel_with_sheets(part2_with_water, "培养基数据2_扩展.xlsx")