import random
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# 定义培养基成分及其取值范围
components = {
    "酵母浸粉": (40, 100),
    "蛋白胨": (40, 100),
    "葡萄糖": (40, 100),
    "(NH4)2SO4": (40, 100),
    "K2HPO4·3H2O": (40, 100),
    "柠檬酸·H2O": (40, 100),
    "硫酸镁": (40, 100),
    "微量元素": (40, 100)
}

# 随机生成96组培养基数据，按照新的取值间隔规则
def generate_random_media(num_samples=96):
    data = []
    for i in range(1, num_samples + 1):
        sample = {"组编号": i}
        for component, (min_val, max_val) in components.items():
            if max_val - min_val > 5:  # 最大范围超过5
                # 以10为间隔生成随机值
                value = random.choice([x for x in range(min_val, max_val + 1, 10)])
            else:  # 最大范围小于等于5
                # 以0.25为间隔生成随机值
                value = round(random.choice([x * 0.25 for x in range(int(min_val * 4), int(max_val * 4) + 1)]), 2)
            sample[component] = value
        data.append(sample)
    return data

# 对数据进行排序并更新组编号
def sort_and_update_group_numbers(data):
    # 将数据转换为DataFrame以便排序
    df = pd.DataFrame(data)
    
    # 按照所有列依次排序
    sorted_df = df.sort_values(by=list(components.keys()))
    
    # 更新组编号
    sorted_df["组编号"] = range(1, len(sorted_df) + 1)
    
    # 返回排序后的数据
    return sorted_df.to_dict(orient="records")

# 检查文件是否存在，若存在则备份
def backup_existing_file(file_name):
    if os.path.exists(file_name):
        # 获取当前时间戳
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 创建备份文件名
        backup_file_name = file_name.replace(".xlsx", f"_backup_{timestamp}.xlsx")
        # 备份原文件
        os.rename(file_name, backup_file_name)
        print(f"原文件已备份为 {backup_file_name}")

# 将数据保存到Excel文件，并美化Sheet2
def save_to_excel_with_sheets(data, file_name="培养基数据.xlsx"):
    # 检查文件是否存在并备份
    backup_existing_file(file_name)
    
    # 创建ExcelWriter对象，用于写入多个sheet
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        # 写入原始数据到sheet1
        df_original = pd.DataFrame(data)
        df_original.to_excel(writer, sheet_name="Sheet1", index=False)
        
        # 创建一个空的DataFrame，用于存储所有成分分布表
        all_tables = []
        
        # 生成所有成分的分布表
        for idx, component in enumerate(components.keys(), start=1):
            # 添加标题行
            title_row = pd.DataFrame([f"第{idx}种成分添加量分布表"], columns=["Title"])
            all_tables.append(title_row)
            
            # 创建一个8行12列的空表格
            rows = ["A", "B", "C", "D", "E", "F", "G", "H"]
            cols = [str(i) for i in range(1, 13)]
            table = pd.DataFrame(index=rows, columns=cols)
            
            # 填充表格数据
            for row_name in rows:
                for col_name in cols:
                    group_number = f"{row_name}{col_name}"  # 组编号
                    # 查找对应组编号的成分值
                    sample = next((s for s in data if s["组编号"] == group_number), None)
                    if sample:
                        table.at[row_name, col_name] = sample[component]
            
            # 将表格添加到all_tables
            all_tables.append(table)
            
            # 添加一个空白行作为分隔
            blank_row = pd.DataFrame([None], columns=["Blank"])
            all_tables.append(blank_row)
        
        # 生成补水量分布表
        title_row_water = pd.DataFrame(["补水量分布表"], columns=["Title"])
        all_tables.append(title_row_water)
        
        # 创建一个8行12列的空表格
        water_table = pd.DataFrame(index=["A", "B", "C", "D", "E", "F", "G", "H"], columns=[str(i) for i in range(1, 13)])
        
        # 填充补水量表格数据
        for row_name in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            for col_name in [str(i) for i in range(1, 13)]:
                group_number = f"{row_name}{col_name}"  # 组编号
                # 查找对应组编号的补水量值
                sample = next((s for s in data if s["组编号"] == group_number), None)
                if sample:
                    water_table.at[row_name, col_name] = sample["补水量"]
        
        # 将补水量表格添加到all_tables
        all_tables.append(water_table)
        
        # 添加一个空白行作为分隔
        blank_row_water = pd.DataFrame([None], columns=["Blank"])
        all_tables.append(blank_row_water)
        
        # 合并所有表格
        combined_table = pd.concat(all_tables, ignore_index=True)
        
        # 写入到Sheet2
        combined_table.to_excel(writer, sheet_name="Sheet2", index=False)
    
    # 打开Excel文件并美化Sheet2
    from openpyxl import load_workbook
    wb = load_workbook(file_name)
    ws = wb["Sheet2"]
    
    # 定义表格起始位置
    current_row = 2
    column = 2   
    # 遍历所有表格并应用样式
    for idx, component in enumerate(list(components.keys()) + ["补水量"], start=2):
        # 标题行样式
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
        current_row += 1
        column = 2         
        # 表格内容样式
        for r in range(8):  # 8行
            for c in range(13):  # 12列
                cell = ws.cell(row=current_row + r, column=c + 2)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center")
        
        # 应用条件格式（覆盖完整8行12列）
        ws.conditional_formatting.add(
            f"A{current_row}:L{current_row + 8}",
            ColorScaleRule(start_type="min", start_color="FF92D050",
                           mid_type="percentile", mid_value=50, mid_color="FFFFC000",
                           end_type="max", end_color="FFFF0000")
        )
        
        # 更新当前行位置
        current_row += 8  # 表格内容
        current_row += 1  # 空白分隔行
    
    # 保存美化后的Excel文件
    wb.save(file_name)
    print(f"数据已成功保存并美化到 {file_name}")

# 计算补水量
def calculate_water_supplement(data):
    num_components = len(components)  # 成分种类数
    for sample in data:
        # 计算该行所有成分值之和
        total_component_value = sum(sample[component] for component in components)
        # 计算补水量
        water_supplement = num_components * 100 - total_component_value + 100
        # 添加补水量列
        sample["补水量"] = water_supplement
    return data

# 重新编号规则
def reassign_group_numbers(data):
    new_data = []
    rows_per_column = 12  # 每列12个编号
    columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
    
    for i, sample in enumerate(data):
        # 计算行号和列号
        column_index = i // rows_per_column
        row_index = i % rows_per_column + 1
        
        # 构造新的编号
        new_sample = sample.copy()
        new_sample["组编号"] = f"{columns[column_index]}{row_index}"
        new_data.append(new_sample)
    
    return new_data

# 新增功能：将48行数据复制为96行数据，并重新编号
def duplicate_and_renumber(data):
    duplicated_data = []
    rows_per_column = 12  # 每列12个编号
    columns = ["A", "B", "C", "D", "E", "F", "G", "H"]
    
    # 将48行数据复制为96行数据，每行重复一次
    for i in range(len(data)):
        duplicated_data.append(data[i])  # 添加原始行
        duplicated_data.append(data[i])  # 再次添加原始行（复制）
    
    # 重新编号
    new_data = []
    for i, sample in enumerate(duplicated_data):
        # 计算行号和列号
        column_index = i // rows_per_column
        row_index = i % rows_per_column + 1
        
        # 构造新的编号
        new_sample = sample.copy()
        new_sample["组编号"] = f"{columns[column_index]}{row_index}"
        new_data.append(new_sample)
    
    return new_data

# 主函数
if __name__ == "__main__":
    # 生成100组随机培养基数据
    media_data = generate_random_media(96)
    
    # 对数据进行排序并更新组编号
    sorted_media_data = sort_and_update_group_numbers(media_data)
    
    # 将96组数据分为两部分，每部分48组
    part1 = sorted_media_data[:48]
    part2 = sorted_media_data[48:]
    
    # 重新编号
    part1_renumbered = reassign_group_numbers(part1)
    part2_renumbered = reassign_group_numbers(part2)
    
    # 将48行数据复制为96行数据，并重新编号
    part1_duplicated = duplicate_and_renumber(part1_renumbered)
    part2_duplicated = duplicate_and_renumber(part2_renumbered)
    
    # 计算补水量
    part1_with_water = calculate_water_supplement(part1_duplicated)
    part2_with_water = calculate_water_supplement(part2_duplicated)
    
    # 保存到两个Excel文件
    save_to_excel_with_sheets(part1_with_water, "培养基数据1_扩展.xlsx")
    save_to_excel_with_sheets(part2_with_water, "培养基数据2_扩展.xlsx")